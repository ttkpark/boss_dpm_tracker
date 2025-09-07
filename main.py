import cv2
import os
from openpyxl import load_workbook
import easyocr
reader = easyocr.Reader(['en'], gpu=True)

VIDEO_PATH = 'video/video9.mp4'
HP_REGION_ORIGINAL = (288, 40, 22, 12)  # (x, y, w, h) → 수정 필요
#HP_REGION_ORIGINAL = (262, 44, 22, 12)  # (x, y, w, h) → 수정 필요
HP_REGION = (0,0,0,0)
SIZE = (0,0)
X_mutli = 1376
Y_mutli = 776
FRAME_INTERVAL_SEC = 0.75  # 1초 간격 # 0.75
tick_max = 4.9  #100 #4.9 #0.9
result_tick = 3.0


def extract_frames(video_path, interval_sec=1):
    global HP_REGION
    global HP_REGION_ORIGINAL
    global X_mutli
    global Y_mutli
    global SIZE

    cap = cv2.VideoCapture(video_path)
    fps = cap.get(cv2.CAP_PROP_FPS)
    interval = int(fps * interval_sec)

    width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))   # ▶ 프레임 너비 (가로)
    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT)) # ▶ 프레임 높이 (세로)
    SIZE = (width, height)
    print(f"영상 해상도: {width} x {height}")

    X_mutli = X_mutli/width
    Y_mutli = Y_mutli/height

    HP_REGION = (int(round(HP_REGION_ORIGINAL[0]/X_mutli,0)), int(round(HP_REGION_ORIGINAL[1]/Y_mutli,0))
              , int(round(HP_REGION_ORIGINAL[2]/X_mutli,0)), int(round(HP_REGION_ORIGINAL[3]/Y_mutli,0)))

    frames, timestamps = [], []
    frame_idx = 0

    while True:
        ret, frame = cap.read()
        if not ret:
            break
        if frame_idx % interval == 0:
            frames.append(frame)
            timestamps.append(frame_idx / fps)
        frame_idx += 1
    cap.release()
    return frames, timestamps

def crop_hp_area(frame, region):
    x, y, w, h = region
    return frame[y:y+h, x:x+w]

def preprocess_hp_image(image, scale=9.0, t=0,thres=0):
    t = int(t*10)
    os.makedirs("debug", exist_ok=True)

    resized = cv2.resize(image, None, fx=scale, fy=scale, interpolation=cv2.INTER_LINEAR)

    gray = cv2.cvtColor(resized, cv2.COLOR_BGR2GRAY)

    blurred = gray  # 블러 제거

    # 이진화
    #thres = 131
    _, thresh = cv2.threshold(blurred, thres, 255, cv2.THRESH_BINARY)

    # morphology
    #kernel1 = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (1, 1))
    #eroded = cv2.erode(thresh, kernel1, iterations=1)

    #kernel2 = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
    #morph = cv2.morphologyEx(eroded, cv2.MORPH_CLOSE, kernel2, iterations=1)

    # 선을 얇게 만들기 위한 추가 erosion
    #kernel3 = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 1))
    final = thresh #cv2.erode(morph, kernel3, iterations=1)

    # OCR 인식을 위해 반전 (흰 글자 / 검정 배경)
    final = cv2.bitwise_not(final)
    #cv2.imwrite(f"debug/{t}_4_morph.png", morph)

    if(t<-2):
        cv2.imwrite(f"debug/{t}_0_resized.png", resized)
        cv2.imwrite(f"debug/{t}_1_gray.png", gray)
        cv2.imwrite(f"debug/{t}_2_blur.png", blurred)
        cv2.imwrite(f"debug/{t}_3_thresh.png", thresh)
    cv2.imwrite(f"debug/{t}_5_final.png", final)

    return final

def read_hp_easyocr(image):
    results = reader.readtext(image, detail=0)
    for text in results:
        if any(char.isdigit() for char in text):
            return text
    return None

def extract_hp_percent(image, t,thres):
    clean = preprocess_hp_image(image,9.0,t,thres)

    res = read_hp_easyocr(clean)
    if(res != None):
        res = res.replace(',', '.').replace('S', '5').replace('s', '5').replace('l', '1').replace('O', '0').replace('o', '0')
        num_of_periods = res.count('.')
        if(num_of_periods > 1):
            first_dot = res.find('.')
            res = res[:first_dot + 1] + res[first_dot + 1:].replace('.', '')
        if(num_of_periods == 0 and not (res == "100" and t < 20)):
            res = res[:-1] + '.' + res[-1]
        try:
            return float(res)
        except ValueError:
            return None


    return None


def save_to_excel(data, filename="boss_hp_log.xlsx"):
    wb = load_workbook("boss_hp_log - template.xlsx")
    ws = wb.active

    for i in range(len(data)):
        ws.cell(row=i+2, column=1).value = data[i][0]  # 1번째 열
        ws.cell(row=i+2, column=2).value = data[i][1]  # 2번째 열

    # 5. 기존 차트 가져오기
    #chart = ws._charts[0]  # 첫 번째 차트만 조작

    # 4. 새로운 차트 범위 재설정
    from openpyxl.chart import ScatterChart, Reference, Series
    chart = ScatterChart()
    chart.title = "구간 DPM 분석"
    #chart.style = 13
    chart.y_axis.title = "딜량(조)"
    chart.x_axis.title = "Time"

    # 데이터 및 카테고리 범위 지정
    row_count = len(data)
    cats = Reference(ws, min_col=3, min_row=2, max_row=row_count+1)  # C열: Time
    
    data = Reference(ws, min_col=6, min_row=1, max_row=row_count+1)  # F열: DPM
    series = Series(data, cats, title_from_data=True)
    chart.series.append(series)

    data = Reference(ws, min_col=7, min_row=1, max_row=row_count+1)  # G열: DPM
    series = Series(data, cats, title_from_data=True)
    chart.series.append(series)

    #chart.series[0].xvalues=cats
    #chart.series[0].values=data
    #chart.series[0].yvalues=data


    # 5. 차트 시트에 추가
    ws.add_chart(chart, "L36")
    wb.save(filename)

    print(f"✅ 저장 완료: {filename}")

def main():
    print("▶ 영상 분석 시작")
    frames, timestamps = extract_frames(VIDEO_PATH, FRAME_INTERVAL_SEC)

    results = []  # (time, hp)
    global X_mutli
    global SIZE
    threses = []
    if(SIZE[0] == 1376):
        threses.append(133)
    elif(SIZE[0] == 1920):
        threses.append(133)
    elif(SIZE[0] == 1364):
        threses.append(160)
    else:
        threses.append(130)
    for thres in threses: #130
        print(f"thres:{thres}")
        results = []
        error_total = 0

        last_hp = None
        global tick_max
        for frame, t in zip(frames, timestamps):
            cropped = crop_hp_area(frame, HP_REGION)
            hp = extract_hp_percent(cropped,t,thres)
            if(last_hp == None and hp != None):
                last_hp = hp
            else:
                if(hp == None or (last_hp - hp) > tick_max or last_hp - hp < 0):
                    delta = 0
                    vw = None
                    if(hp != None):
                        delta = last_hp - hp
                        vw = hp
                    print(f"{t:.1f}s: {vw}% Invalid delta = {delta}%")
                    error_total = error_total+1
                    continue

                if(last_hp == hp):
                    continue

                last_hp = hp
            
            print(f"{t:.1f}s: {hp}%")
            results.append((round(t, 1), str(round(hp, 1))+"%"))

        print(f"thres = {thres}, last% = {last_hp}, fail_count = {error_total}")

    global result_tick

    # 5초마다 딜량 뽑기
    deals_5sec = []
    focus_sec = 0
    last_x = results[0]
    for x in results:
        x_sec = x[0]
        if(focus_sec < x_sec):
            deals_5sec.append((focus_sec,last_x[1]))
            print(f"{focus_sec}s: {last_x[1]}")
            focus_sec += result_tick

        last_x = x

    # 2. 엑셀로 저장
    save_to_excel(deals_5sec,f"boss_hp_log_{VIDEO_PATH.split('/')[-1]}.xlsx")

if __name__ == "__main__":
    main()